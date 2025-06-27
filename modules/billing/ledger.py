from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from datetime import datetime

LEDGER_PATH = "data/ledger.xlsx"

def ensure_workbook():
    if not os.path.exists(LEDGER_PATH):
        wb = Workbook()
        wb.remove(wb.active)

        wb.create_sheet("Безнал")  # Страница 1
        wb.create_sheet("Наличные")  # Страница 2

        for sheet in wb.worksheets:
            sheet.append(["User ID", "Файл", "Страниц", "Цена", "Время"])
            for col in range(1, 6):
                sheet.column_dimensions[get_column_letter(col)].width = 20

        wb.save(LEDGER_PATH)

def log_print_job(user_id: int, file_name: str, page_count: int, price: int, method: str):
    ensure_workbook()
    wb = load_workbook(LEDGER_PATH)

    sheet_name = "Безнал" if method == "card" else "Наличные"
    sheet = wb[sheet_name]

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([user_id, file_name, page_count, price, timestamp])

    wb.save(LEDGER_PATH)
